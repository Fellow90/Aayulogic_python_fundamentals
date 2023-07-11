
'''
import requests, json
from openpyxl import Workbook, load_workbook

url = 'https://gist.githubusercontent.com/saniyusuf/406b843afdfb9c6a86e25753fe2761f4/raw/523c324c7fcc36efab8224f9ebb7556c09b69a14/Film.JSON'
response = requests.get(url)
films = response.json()
# for i in films:
#     print("\n",i)

headers = []
for i in films[0].keys():
    headers.append(i.capitalize())
print(headers)  

wb = Workbook()
ws = wb.active
ws.append(headers)
for row in range(1,14):
    ws.cell(row = row+1, column = 1, value = 'False')
# ws.title('Here comes the details of the movies')
for film in films:
    values = []
    for key in headers:
        if key == "Comingsoon":
            values.append('False')
        else:
            values.append(film.get(key," "))
    ws.append(values)
    # ws.append(list(film.values()))
wb.save('output.xlsx')
'''

'''
import requests,json
from openpyxl import Workbook

url = 'https://raw.githubusercontent.com/benoitvallon/100-best-books/master/books.json'
response = requests.get(url)
if response.status_code == 200:
    books = response.json()
else:
    print("Failed to fetch data from the URL")
for i in books:
    print("\n",i)
from openpyxl.styles import Font
booksKeys = []
for key in books[0].keys():
    booksKeys.append(key.capitalize())

wb = Workbook()
sheet = wb.active
sheet.append(booksKeys)
fo = Font(italic=True, bold= True, size=20)
for cell in sheet["1:1"]:
    cell.font = fo
for each in books:
    if each['author'] != "Unknown":  
        sheet.append(list(each.values()))
wb.save('output.xlsx')'''
'''
import requests,json
from openpyxl import Workbook
url = 'https://raw.githubusercontent.com/benoitvallon/100-best-books/master/books.json'
response = requests.get(url)
if response.status_code == 200:
    books = response.json()
else:
    print("Failed to load url")

def filterkeybyvalue(yourkey):
    userinput = input("Enter the key to search on the dictionary:")
    booksdictionary = {}
    for book in books():
        author = book['author']
        country = book['country']
        imagelink = book['imagelink']
        language = book['language']
        link = book['link']
        page = book['page']
        title = book['title']
        year = book['year']
        if yourkey not in books[0].keys():
            if yourkey not in booksdictionary:
                booksdictionary[yourkey] = [book]
            else:
                booksdictionary[yourkey].append(book)
            
        else:
            print("Please enter the valid keys. We have got details about: ",books[0].keys())
    return booksdictionary

print(filterkeybyvalue())'''

import requests,json
from openpyxl import Workbook
url = 'https://raw.githubusercontent.com/benoitvallon/100-best-books/master/books.json'
response = requests.get(url)
if response.status_code == 200:
    books = response.json()
else:
    print("Failed to load url")
'''
def filterkeylist(listdict,key,value):
    
    listofdictionary = {}
    
    for dictionary in listdict:
        if key in dictionary and dictionary[key] == value:
                if key not in listofdictionary:
                    listofdictionary[key] = [dictionary]
                else:
                    listofdictionary[key].append(dictionary)
        # else:
        #     print("We dont have any value like you provided.")

        else:
            print("Please enter the valid keys. We have got the details regarding ", list(dictionary.keys()))
            

    return listofdictionary

    
keyinput = input("Please enter the key to search: ")
valueinput = input("Which value of the key you want to search for: ")
print(filterkeylist(books,keyinput, valueinput))'''

'''def filterkeylist(listdict,key,value):
    listofdictionary = {}
    for dictionary in listdict:
        if key not in dictionary:
            print("Please enter the valid keys. We have got the details regarding ", list(dictionary.keys()))
        else:            
            if dictionary[key] != value:
                print("We dont have any value like you provided.")
                break
            else:
                if key not in listofdictionary:
                    listofdictionary[key] = [dictionary]
                else:
                    listofdictionary[key].append(dictionary)
    return listofdictionary

    
keyinput = input("Please enter the key to search: ")
valueinput = input("Which value of the key you want to search for: ")
print(filterkeylist(books,keyinput, valueinput))'''

   




'''def filterkeylist(listdict,key,value):
    listofdictionary = []
    for dictionary in listdict:
        if key in dictionary:
            if dictionary[key] == value:    
                listofdictionary.append(dictionary)
        if key not in dictionary:
            print(f"We dont have the information regarding the key you provided.\nWe only have:{list(dictionary.keys())} ")
            break

                
    return listofdictionary

    
keyinput = input("Please enter the key to search: ")
valueinput = input("Which value of the key you want to search for: ")
print(filterkeylist(books,keyinput, valueinput))'''


'''keyinput = input("Please enter the key to search: ")
valueinput = input("Which value of the key do you want to search for: ")

def filterkeylist(key, value):
    listofdictionary = []
    for dictionary in books:
        if key in dictionary and str(dictionary[key]).lower() == str(value).lower():
            listofdictionary.append(dictionary)
        
        if key not in dictionary:
            print(f"We don't have the information regarding the key you provided.\nWe only have: {list(dictionary.keys())}")

    return listofdictionary
    '''

'''def datatypeconverter():
    valueinput = input("Which value of the key do you want to search for: ")
    valueinput2 = int(input("Which value of the key do you want to search for: "))
    return valueinput, valueinput2

a,b = datatypeconverter()
'''




'''

print(filterkeylist(keyinput, valueinput))
'''



keyinput = input("Please enter the key to search: ")
valueinput = input("Which value of the key do you want to search for: ")

def filterkeylist(key, value):
    listofdictionary = []
    for dictionary in books:
        # if isinstance(value, str) and isinstance(dictionary[key],str) and value.startswith("http://") or value.startswith("https://"):
        #     if str(dictionary[key]).lower() == str(value).lower():
        #         listofdictionary.append(dictionary)

        if key in dictionary and str(dictionary[key]).lower() == str(value).lower():
            listofdictionary.append(dictionary)
        
        if key not in dictionary:
            return print(f"We don't have the information regarding the key you provided.\nWe only have: {list(dictionary.keys())}")
            

            

    return listofdictionary

print(filterkeylist(keyinput,valueinput))